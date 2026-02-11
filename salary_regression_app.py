"""
薪酬回归分析工具 - Streamlit云端版
GitHub+Streamlit部署，支持文件上传/在线分析/结果下载
核心算法保留原对数变换回归逻辑
"""
import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
import warnings
from io import BytesIO  # 内存中生成Excel，支持下载
warnings.filterwarnings('ignore')

# ---------------------- 保留原核心回归算法类 ----------------------
class SalaryRegressionImproved:
    """改进的薪酬回归分析类 - 使用对数变换（核心算法无改动）"""
    def __init__(self):
        self.input_data = None
        self.params = {}
        self.models = {}
        self.formulas = {}
        self.results = None
        self.metrics = None

    def log_polynomial_regression(self, X, y, degree=2):
        """对数多项式回归 - 原逻辑完全保留"""
        valid_mask = (~np.isnan(y)) & (y > 0)
        X_valid = X[valid_mask].reshape(-1, 1)
        y_valid = y[valid_mask]
        if len(X_valid) < 3:
            return None, None, None
        log_y_valid = np.log(y_valid)
        poly = PolynomialFeatures(degree=degree)
        X_poly = poly.fit_transform(X_valid)
        model = LinearRegression()
        model.fit(X_poly, log_y_valid)
        return model, poly, y_valid

    def get_formula_string(self, model, poly, percentile):
        """生成回归公式字符串 - 原逻辑完全保留"""
        degree = self.params['poly_degree']
        intercept = model.intercept_
        coefs = model.coef_[1:]
        formula_parts_log = [f"{intercept:.6f}"]
        for i, coef in enumerate(coefs):
            power = i + 1
            sign = "+" if coef >= 0 else ""
            if power == 1:
                formula_parts_log.append(f"{sign}{coef:.6f}*x")
            else:
                formula_parts_log.append(f"{sign}{coef:.6f}*x^{power}")
        log_formula = " ".join(formula_parts_log)
        A = np.exp(intercept)
        if degree == 1:
            b = coefs[0]
            formula = f"{A:.2f} * exp({b:.6f}*x)"
        elif degree == 2:
            b, c = coefs[0], coefs[1]
            formula = f"{A:.2f} * exp({b:.6f}*x + {c:.6f}*x^2)"
        else:
            formula = f"exp({log_formula})"
        self.formulas[percentile] = {
            'log_formula': log_formula,
            'formula': formula,
            'intercept': intercept,
            'coefficients': coefs.tolist(),
            'degree': degree,
            'A': A
        }
        return formula, log_formula

    def predict_percentiles(self):
        """对各分位数进行回归预测 - 原逻辑保留，仅移除打印"""
        grades = self.input_data['Survey Grade'].values
        percentiles = ['P10', 'P25', 'P50', 'P75', 'P90']
        target_grades = np.arange(
            self.params['grade_start'],
            self.params['grade_end'] + self.params['grade_step'],
            self.params['grade_step']
        )
        results = pd.DataFrame({'Survey Grade': target_grades})
        results = results.sort_values('Survey Grade', ascending=False)
        for percentile in percentiles:
            if percentile not in self.input_data.columns:
                continue
            y = self.input_data[percentile].values
            model, poly, y_train = self.log_polynomial_regression(grades, y, degree=self.params['poly_degree'])
            if model is None:
                continue
            self.models[percentile] = {'model': model, 'poly': poly, 'y_train': y_train}
            formula, log_formula = self.get_formula_string(model, poly, percentile)
            X_target = poly.transform(target_grades.reshape(-1, 1))
            log_y_pred = model.predict(X_target)
            y_pred = np.exp(log_y_pred)
            grade_to_pred = dict(zip(target_grades, y_pred))
            results[percentile] = results['Survey Grade'].map(grade_to_pred)
        self.results = results
        return results

    def calculate_metrics(self):
        """计算回归质量指标 - 原逻辑保留，仅移除打印"""
        metrics = []
        percentiles = ['P10', 'P25', 'P50', 'P75', 'P90']
        grades = self.input_data['Survey Grade'].values
        for percentile in percentiles:
            if percentile not in self.input_data.columns or percentile not in self.models:
                continue
            y_original = self.input_data[percentile].values
            valid_mask = (~np.isnan(y_original)) & (y_original > 0)
            if valid_mask.sum() == 0:
                continue
            model_info = self.models[percentile]
            model = model_info['model']
            poly = model_info['poly']
            X_original = poly.transform(grades[valid_mask].reshape(-1, 1))
            log_y_pred = model.predict(X_original)
            y_pred = np.exp(log_y_pred)
            y_actual = y_original[valid_mask]
            ss_res = np.sum((y_actual - y_pred) ** 2)
            ss_tot = np.sum((y_actual - y_actual.mean()) ** 2)
            r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0
            mape = np.mean(np.abs((y_actual - y_pred) / y_actual)) * 100
            metrics.append({
                '分位数': percentile,
                'R2': f'{r_squared:.4f}',
                '平均误差%': f'{mape:.2f}%',
                '样本数': int(valid_mask.sum()),
                '回归公式': self.formulas[percentile]['formula']
            })
        self.metrics = pd.DataFrame(metrics)
        return self.metrics

    def generate_excel_report(self):
        """【改造核心】生成Excel报告到内存（代替本地保存），支持Streamlit下载"""
        output = BytesIO()  # 内存文件对象
        wb = openpyxl.Workbook()  # 新建工作簿（代替读取本地）
        # 删除默认sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        # 保存各结果sheet（原逻辑保留，仅修改wb为新建的工作簿）
        self._save_regression_summary(wb)
        self._save_formulas(wb)
        self._save_metrics(wb)
        self._save_comparison(wb)
        self._add_summary_chart(wb)
        # 保存到内存
        wb.save(output)
        output.seek(0)  # 指针回到开头，方便下载
        return output

    # 以下为原保存逻辑的封装，仅修改为接收外部wb，移除本地文件依赖
    def _save_regression_summary(self, wb):
        ws = wb.create_sheet('回归结果汇总', 0)
        headers = list(self.results.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx, row_data in enumerate(self.results.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row_idx, col_idx, value)
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.number_format = '#,##0.00'
                    if col_idx == 2:
                        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    elif col_idx == 3:
                        cell.fill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
                    elif col_idx == 4:
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    elif col_idx == 5:
                        cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                    elif col_idx == 6:
                        cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

    def _save_comparison(self, wb):
        ws = wb.create_sheet('原始vs回归对比')
        ws.cell(1, 1, '原始数据 vs 回归结果对比')
        ws.cell(1, 1).font = Font(size=14, bold=True, color='FFFFFF')
        ws.cell(1, 1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.merge_cells('A1:L1')
        headers = ['职级', 'P50原始', 'P50回归', '差异', '差异%', 'P25原始', 'P25回归', 'P75原始', 'P75回归', '职级增长', 'P50增长率', 'P50增长额']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(3, col_idx, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row_idx = 4
        prev_p50_regressed = None
        prev_grade = None
        for _, result_row in self.results.iterrows():
            grade = result_row['Survey Grade']
            p50_regressed = result_row['P50']
            p25_regressed = result_row['P25'] if 'P25' in result_row else None
            p75_regressed = result_row['P75'] if 'P75' in result_row else None
            original_row = self.input_data[self.input_data['Survey Grade'] == grade]
            if not original_row.empty:
                p50_original = original_row['P50'].values[0]
                p25_original = original_row['P25'].values[0] if 'P25' in original_row else None
                p75_original = original_row['P75'].values[0] if 'P75' in original_row else None
                diff = p50_regressed - p50_original
                diff_pct = (diff / p50_original * 100) if p50_original > 0 else 0
            else:
                p50_original = None
                p25_original = None
                p75_original = None
                diff = None
                diff_pct = None
            if prev_p50_regressed is not None and prev_grade is not None:
                growth_rate = (prev_p50_regressed - p50_regressed) / p50_regressed * 100
                growth_amount = prev_p50_regressed - p50_regressed
                grade_change = f"{grade:.0f}->{prev_grade:.0f}"
            else:
                growth_rate = None
                growth_amount = None
                grade_change = ""
            ws.cell(row_idx, 1, grade)
            ws.cell(row_idx, 2, p50_original)
            ws.cell(row_idx, 3, p50_regressed)
            ws.cell(row_idx, 4, diff)
            ws.cell(row_idx, 5, diff_pct)
            ws.cell(row_idx, 6, p25_original)
            ws.cell(row_idx, 7, p25_regressed)
            ws.cell(row_idx, 8, p75_original)
            ws.cell(row_idx, 9, p75_regressed)
            ws.cell(row_idx, 10, grade_change)
            ws.cell(row_idx, 11, growth_rate)
            ws.cell(row_idx, 12, growth_amount)
            for col in [2, 3, 4, 6, 7, 8, 9, 12]:
                if ws.cell(row_idx, col).value is not None:
                    ws.cell(row_idx, col).number_format = '#,##0'
            for col in [5, 11]:
                if ws.cell(row_idx, col).value is not None:
                    ws.cell(row_idx, col).number_format = '0.0"%"'
            row_idx += 1
            prev_p50_regressed = p50_regressed
            prev_grade = grade
        ws.column_dimensions['A'].width = 8
        for col in ['B', 'C', 'D', 'F', 'G', 'H', 'I', 'L']:
            ws.column_dimensions[col].width = 13
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12

    def _save_formulas(self, wb):
        ws = wb.create_sheet('回归公式')
        ws.cell(1, 1, '回归公式详情（对数变换回归）')
        ws.cell(1, 1).font = Font(size=14, bold=True, color='FFFFFF')
        ws.cell(1, 1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws.merge_cells('A1:D1')
        headers = ['分位数', '回归公式', '多项式阶数', 'Excel公式（近似）']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(3, col_idx, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        row_idx = 4
        for percentile, info in self.formulas.items():
            ws.cell(row_idx, 1, percentile)
            ws.cell(row_idx, 2, f"y = {info['formula']}")
            ws.cell(row_idx, 3, info['degree'])
            A = info['A']
            coefs = info['coefficients']
            if len(coefs) >= 2:
                b, c = coefs[0], coefs[1]
                excel_formula = f"={A:.2f}*EXP({b:.6f}*x+{c:.6f}*POWER(x,2))"
            else:
                b = coefs[0]
                excel_formula = f"={A:.2f}*EXP({b:.6f}*x)"
            ws.cell(row_idx, 4, excel_formula)
            row_idx += 1
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        instructions = [
            '1. 本工具使用对数变换回归，适合薪酬指数增长特性',
            '2. 回归公式：y = A * exp(b*x + c*x^2)，其中x为职级',
            '3. Excel公式：将x替换为单元格引用（如A2）',
            '4. 优势：避免低职级曲线过平、高职级过陡的问题',
            f'5. 当前使用{self.params["poly_degree"]}阶对数多项式回归',
        ]
        ws.cell(row_idx + 2, 1, '使用说明：')
        ws.cell(row_idx + 2, 1).font = Font(bold=True)
        for i, instruction in enumerate(instructions):
            ws.cell(row_idx + 3 + i, 1, instruction)
            ws.merge_cells(f'A{row_idx + 3 + i}:D{row_idx + 3 + i}')

    def _save_metrics(self, wb):
        ws_metrics = wb.create_sheet('回归指标')
        metric_headers = list(self.metrics.columns)
        for col_idx, header in enumerate(metric_headers, 1):
            cell = ws_metrics.cell(1, col_idx, header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        for row_idx, row_data in enumerate(self.metrics.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_metrics.cell(row_idx, col_idx, value)
        ws_metrics.column_dimensions['A'].width = 12
        ws_metrics.column_dimensions['B'].width = 12
        ws_metrics.column_dimensions['C'].width = 12
        ws_metrics.column_dimensions['D'].width = 10
        ws_metrics.column_dimensions['E'].width = 60

    def _add_summary_chart(self, wb):
        ws = wb['回归结果汇总']
        chart = ScatterChart()
        chart.title = "薪酬回归曲线汇总（对数变换回归）"
        chart.x_axis.title = "Survey Grade"
        chart.y_axis.title = "Salary"
        chart.height = 15
        chart.width = 25
        x_values = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        percentiles = ['P10', 'P25', 'P50', 'P75', 'P90']
        for idx, percentile in enumerate(percentiles):
            if percentile in list(self.results.columns):
                col_idx = list(self.results.columns).index(percentile) + 1
                y_values = Reference(ws, min_col=col_idx, min_row=2, max_row=ws.max_row)
                series = Series(y_values, x_values, title=percentile)
                chart.series.append(series)
                chart.series[idx].marker.symbol = "circle"
                chart.series[idx].marker.size = 5
                chart.series[idx].smooth = True
                chart.series[idx].graphicalProperties.line.width = 25000
        ws.add_chart(chart, f'A{ws.max_row + 3}')

# ---------------------- Streamlit前端交互逻辑 ----------------------
def main():
    # 页面配置（标题、图标、布局）
    st.set_page_config(
        page_title="薪酬回归分析工具-云端版",
        page_icon="📊",
        layout="wide"  # 宽屏布局
    )
    # 页面标题
    st.title("📊 薪酬回归分析工具（对数变换改进版）")
    st.divider()  # 分隔线

    # 1. 侧边栏：参数配置
    with st.sidebar:
        st.header("⚙️ 分析参数配置")
        poly_degree = st.selectbox("多项式阶数", options=[1, 2], value=2, help="原工具默认2阶，适合薪酬指数增长")
        grade_start = st.number_input("目标职级起始", min_value=1, max_value=100, value=1, step=1)
        grade_end = st.number_input("目标职级结束", min_value=grade_start, max_value=100, value=20, step=1)
        grade_step = st.number_input("职级步长", min_value=0.5, max_value=5.0, value=1.0, step=0.5)
        st.info("参数说明：与原Excel工具参数一致，建议保持默认", icon="ℹ️")

    # 2. 主区域：文件上传
    st.subheader("📁 上传Excel数据文件")
    st.caption("要求：Excel需包含【数据输入】sheet，列必须有Survey Grade、P10、P25、P50、P75、P90")
    uploaded_file = st.file_uploader("选择Excel文件", type=["xlsx"], accept_multiple_files=False)

    # 3. 核心流程：上传文件后运行分析
    if uploaded_file is not None:
        try:
            # 读取上传的Excel文件
            df_input = pd.read_excel(uploaded_file, sheet_name='数据输入')
            # 数据校验
            required_cols = ['Survey Grade', 'P10', 'P25', 'P50', 'P75', 'P90']
            if not all(col in df_input.columns for col in required_cols):
                st.error(f"Excel文件缺少必要列！必须包含：{required_cols}")
                return
            df_input = df_input.dropna(subset=['Survey Grade'])
            st.success(f"✅ 成功读取数据：{len(df_input)}行有效数据")
            st.dataframe(df_input.head(10), use_container_width=True)  # 预览数据

            # 初始化分析类
            reg = SalaryRegressionImproved()
            reg.input_data = df_input
            # 设置参数（从侧边栏获取）
            reg.params = {
                'poly_degree': poly_degree,
                'grade_start': grade_start,
                'grade_end': grade_end,
                'grade_step': grade_step,
                'salary_item': '薪酬'
            }

            # 运行分析（按钮触发，避免重复计算）
            if st.button("🚀 开始薪酬回归分析", type="primary"):
                with st.spinner("正在进行对数多项式回归分析..."):
                    # 执行核心分析
                    reg.predict_percentiles()
                    reg.calculate_metrics()
                    # 生成Excel报告
                    excel_output = reg.generate_excel_report()

                # 4. 展示分析结果
                st.divider()
                st.subheader("📈 回归分析结果展示")
                # 分栏展示核心结果
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("### 📊 回归结果汇总")
                    st.dataframe(reg.results, use_container_width=True, format_func=lambda x: f"{x:,.2f}" if isinstance(x, (int, float)) else x)
                with col2:
                    st.markdown("### 📋 回归质量指标")
                    st.dataframe(reg.metrics, use_container_width=True)

                # 展示回归公式
                st.markdown("### 📝 回归公式详情")
                formula_df = pd.DataFrame([
                    {'分位数': k, '回归公式': v['formula'], 'Excel近似公式': f"={v['A']:.2f}*EXP({v['coefficients'][0]:.6f}*x{'+'+str(v['coefficients'][1]):.6f}*POWER(x,2))" if len(v['coefficients'])>=2 else f"={v['A']:.2f}*EXP({v['coefficients'][0]:.6f}*x)"}
                    for k, v in reg.formulas.items()
                ])
                st.dataframe(formula_df, use_container_width=True)

                # 5. 提供Excel报告下载
                st.divider()
                st.subheader("📥 下载完整分析报告")
                st.download_button(
                    label="📄 下载Excel分析报告",
                    data=excel_output,
                    file_name=f"薪酬回归分析报告_{pd.Timestamp.now().strftime('%Y%m%d%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                st.success("✅ 分析完成！点击上方按钮下载完整Excel报告（含所有sheet+图表）")

        except Exception as e:
            st.error(f"❌ 分析失败：{str(e)}", icon="🚨")
            st.exception(e)  # 打印详细错误信息，方便调试

if __name__ == '__main__':
    main()
